import sys
import os

def main():
    if len(sys.argv) < 2:
        print("Usage: enable_gridlines.py <excel_file_path>")
        sys.exit(1)
        
    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"ERROR: File not found at: {filepath}")
        sys.exit(1)
        
    try:
        import openpyxl
        from openpyxl.styles import Border, Side
        
        # Define a high-fidelity light-gray side border representing Excel gridlines
        grid_side = Side(style='thin', color='D3D3D3')
        grid_border = Border(left=grid_side, right=grid_side, top=grid_side, bottom=grid_side)
        
        wb = openpyxl.load_workbook(filepath)
        for sheet in wb.worksheets:
            # 1. Force gridlines visible in editor/view
            try:
                sheet.sheet_view.showGridLines = True
            except:
                pass

            # 2. Force gridlines in print options
            try:
                sheet.print_options.gridLines = True
            except:
                pass
                
            # 3. Apply explicit light-gray borders to all cells in the active range!
            # This is the 100% bulletproof solution that works on LibreOffice Calc and all OS.
            try:
                max_r = sheet.max_row
                max_c = sheet.max_column
                # Only apply if sheet has data
                if max_r > 0 and max_c > 0:
                    for row in sheet.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=max_c):
                        for cell in row:
                            # Apply grid border
                            cell.border = grid_border
            except Exception as e:
                print(f"Warning (borders): {e}")
                
            # 4. Force scaling to fit exactly 1 page wide and 1 page tall
            try:
                if not sheet.sheet_properties.pageSetUpPr:
                    sheet.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
                else:
                    sheet.sheet_properties.pageSetUpPr.fitToPage = True
                
                sheet.page_setup.fitToWidth = 1
                sheet.page_setup.fitToHeight = 1
                
                # Dynamically choose landscape orientation if table is wide (improves fitting!)
                if sheet.max_column > sheet.max_row:
                    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
                else:
                    sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
            except Exception as e:
                print(f"Warning (pageSetup): {e}")

        wb.save(filepath)
        print("SUCCESS")
    except Exception as e:
        print(f"ERROR: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
